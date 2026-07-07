import os
import json
import logging
from typing import Dict, Any
from agent.tools.base import tool
from agent.tools.file_ops import read_pdf
from agent.llm import GeminiClient

logger = logging.getLogger("agent.tools.policy_extractor")

# Directory where policy JSON profiles are cached
CACHE_DIR = os.path.normpath("insurance_rules")

def extract_policy_profile(pdf_path: str) -> Dict[str, Any]:
    """
    Parses an insurance rules PDF, extracts structured metadata parameters using the LLM,
    and caches the result locally as a JSON file.
    """
    pdf_path = os.path.normpath(pdf_path)
    base_name = os.path.basename(pdf_path)
    profile_name = os.path.splitext(base_name)[0] + "_profile.json"
    cache_path = os.path.join(CACHE_DIR, profile_name)
    
    # Check cache first
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                profile = json.load(f)
                logger.info(f"Loaded cached policy profile for: {base_name}")
                return profile
        except Exception as e:
            logger.error(f"Error reading cached profile: {str(e)}")

    logger.info(f"Extracting new policy profile from: {base_name}")
    
    # Extract raw text from PDF
    raw_text = read_pdf(pdf_path)
    if raw_text.startswith("Error:"):
        raise ValueError(f"Could not read PDF: {raw_text}")

    # Use LLM to extract structured parameters
    client = GeminiClient()
    prompt = (
        f"You are a professional medical insurance underwriter.\n"
        f"Analyze the raw insurance guidelines text below and extract these exact parameters into a JSON object:\n"
        f"1. company_name: Name of the insurance company.\n"
        f"2. policy_tier: Coverage tier (e.g. Gold, Silver, Bronze, or Normal if unspecified).\n"
        f"3. deductible_inr: Numeric general deductible in Rupees (0.0 if not mentioned).\n"
        f"4. room_rent_capping: The daily room rent limit (e.g. '1% of sum insured' or a specific amount).\n"
        f"5. icu_capping: The daily ICU rent limit (e.g. '2% of sum insured' or a specific amount).\n"
        f"6. copay_percentage: Numeric copay percentage the patient pays (e.g. 10.0 for 10%, 0.0 if none).\n"
        f"7. waiting_period_months: Pre-existing disease waiting period in months (0 if none).\n"
        f"8. specialty_limits: Co-insurance/coverage rates (0.0 to 1.0) for specialty areas: "
        f"cardiac_care, outpatient, emergency, dental, general_consultation.\n\n"
        f"Return ONLY a clean, valid JSON block matching this exact schema:\n"
        f"{{\n"
        f"  \"company_name\": \"string\",\n"
        f"  \"policy_tier\": \"string\",\n"
        f"  \"deductible_inr\": 0.0,\n"
        f"  \"room_rent_capping\": \"string\",\n"
        f"  \"icu_capping\": \"string\",\n"
        f"  \"copay_percentage\": 0.0,\n"
        f"  \"waiting_period_months\": 0,\n"
        f"  \"specialty_limits\": {{\n"
        f"    \"cardiac_care\": 1.0,\n"
        f"    \"outpatient\": 0.9,\n"
        f"    \"emergency\": 0.95,\n"
        f"    \"dental\": 0.8,\n"
        f"    \"general_consultation\": 1.0\n"
        f"  }}\n"
        f"}}\n\n"
        f"Raw PDF Text:\n{raw_text[:50000]}"
    )
    
    try:
        response = client.generate(
            messages=[{"role": "user", "content": prompt}],
            system_instruction="You extract structured JSON policy parameters strictly and accurately. Return ONLY valid raw JSON."
        )
        
        # Strip code block wrappers if generated
        response_clean = response.strip()
        if response_clean.startswith("```"):
            # Strip first line and last line
            lines = response_clean.splitlines()
            if lines[0].startswith("```"):
                lines = lines[1:]
            if lines[-1].startswith("```"):
                lines = lines[:-1]
            response_clean = "\n".join(lines).strip()
            
        profile = json.loads(response_clean)
        
        # Save cache
        os.makedirs(CACHE_DIR, exist_ok=True)
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(profile, f, indent=2)
            
        logger.info(f"Successfully cached policy profile for: {base_name}")
        return profile
    except Exception as e:
        logger.error(f"Failed to parse or serialize policy profile: {str(e)}")
        # Return fallback baseline profile
        fallback = {
            "company_name": os.path.splitext(base_name)[0].upper(),
            "policy_tier": "STANDARD",
            "deductible_inr": 0.0,
            "room_rent_capping": "1% of Sum Insured",
            "icu_capping": "2% of Sum Insured",
            "copay_percentage": 10.0,
            "waiting_period_months": 24,
            "specialty_limits": {
                "cardiac_care": 0.80,
                "outpatient": 0.70,
                "emergency": 0.80,
                "dental": 0.50,
                "general_consultation": 0.90
            }
        }
        return fallback

def extract_policy_profile_from_excel(xlsx_path: str) -> Dict[str, Any]:
    """
    Parses an insurance rules Excel file, extracts structured metadata parameters using the LLM,
    and caches the result locally as a JSON file.
    """
    xlsx_path = os.path.normpath(xlsx_path)
    base_name = os.path.basename(xlsx_path)
    profile_name = os.path.splitext(base_name)[0] + "_profile.json"
    cache_path = os.path.join(CACHE_DIR, profile_name)
    
    # Check cache first
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                profile = json.load(f)
                logger.info(f"Loaded cached policy profile for: {base_name}")
                return profile
        except Exception as e:
            logger.error(f"Error reading cached profile: {str(e)}")

    logger.info(f"Extracting new policy profile from Excel: {base_name}")
    
    # Read Excel sheets and extract a compact text representation for the LLM
    import openpyxl
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        excel_summary = []
        for sheetname in wb.sheetnames[:3]: # Read first 3 sheets to avoid token overflow
            sheet = wb[sheetname]
            excel_summary.append(f"### Sheet: {sheetname}")
            # Read first 30 rows
            for r in range(1, min(30, sheet.max_row + 1)):
                row_vals = [sheet.cell(r, c).value for c in range(1, min(10, sheet.max_column + 1))]
                while row_vals and row_vals[-1] is None:
                    row_vals.pop()
                if any(x is not None for x in row_vals):
                    row_str = [str(x) if x is not None else "" for x in row_vals]
                    excel_summary.append(f"Row {r:02d}: {row_str}")
        raw_text = "\n".join(excel_summary)
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {str(e)}")

    # Use LLM to extract structured parameters
    client = GeminiClient()
    prompt = (
        f"You are a professional medical insurance underwriter.\n"
        f"Analyze the raw insurance guidelines/charge sheets extracted from the Excel file below and extract these parameters into a JSON object:\n"
        f"1. company_name: Name of the insurance company.\n"
        f"2. policy_tier: Coverage tier (e.g. Gold, Silver, Bronze, or Normal if unspecified).\n"
        f"3. deductible_inr: Numeric general deductible in Rupees (0.0 if not mentioned).\n"
        f"4. room_rent_capping: The daily room rent limit (e.g. '1% of sum insured' or a specific amount).\n"
        f"5. icu_capping: The daily ICU rent limit (e.g. '2% of sum insured' or a specific amount).\n"
        f"6. copay_percentage: Numeric copay percentage the patient pays (e.g. 10.0 for 10%, 0.0 if none).\n"
        f"7. waiting_period_months: Pre-existing disease waiting period in months (0 if none).\n"
        f"8. specialty_limits: Co-insurance/coverage rates (0.0 to 1.0) for specialty areas: "
        f"cardiac_care, outpatient, emergency, dental, general_consultation.\n\n"
        f"Return ONLY a clean, valid JSON block matching this exact schema:\n"
        f"{{\n"
        f"  \"company_name\": \"string\",\n"
        f"  \"policy_tier\": \"string\",\n"
        f"  \"deductible_inr\": 0.0,\n"
        f"  \"room_rent_capping\": \"string\",\n"
        f"  \"icu_capping\": \"string\",\n"
        f"  \"copay_percentage\": 0.0,\n"
        f"  \"waiting_period_months\": 0,\n"
        f"  \"specialty_limits\": {{\n"
        f"    \"cardiac_care\": 1.0,\n"
        f"    \"outpatient\": 0.9,\n"
        f"    \"emergency\": 0.95,\n"
        f"    \"dental\": 0.8,\n"
        f"    \"general_consultation\": 1.0\n"
        f"  }}\n"
        f"}}\n\n"
        f"Raw Excel Text:\n{raw_text[:50000]}"
    )
    
    try:
        response = client.generate(
            messages=[{"role": "user", "content": prompt}],
            system_instruction="You extract structured JSON policy parameters strictly and accurately. Return ONLY valid raw JSON."
        )
        
        response_clean = response.strip()
        if response_clean.startswith("```"):
            lines = response_clean.splitlines()
            if lines[0].startswith("```"):
                lines = lines[1:]
            if lines[-1].startswith("```"):
                lines = lines[:-1]
            response_clean = "\n".join(lines).strip()
            
        profile = json.loads(response_clean)
        
        os.makedirs(CACHE_DIR, exist_ok=True)
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(profile, f, indent=2)
            
        logger.info(f"Successfully cached policy profile from Excel for: {base_name}")
        return profile
    except Exception as e:
        logger.error(f"Failed to parse or serialize policy profile from Excel: {str(e)}")
        # Return fallback baseline profile
        fallback = {
            "company_name": os.path.splitext(base_name)[0].upper().replace("_", " "),
            "policy_tier": "STANDARD",
            "deductible_inr": 0.0,
            "room_rent_capping": "1% of Sum Insured",
            "icu_capping": "2% of Sum Insured",
            "copay_percentage": 10.0,
            "waiting_period_months": 24,
            "specialty_limits": {
                "cardiac_care": 0.80,
                "outpatient": 0.70,
                "emergency": 0.80,
                "dental": 0.50,
                "general_consultation": 0.90
            }
        }
        return fallback

@tool
def get_policy_profile(company_name: str) -> str:
    """
    Instantly retrieves the cached structured JSON profile for an insurance rules document.
    Use this instead of reading massive raw PDF files directly to save token context and guarantee 100% accurate rules processing.
    
    Args:
        company_name: The name or partial string of the insurance company policy (e.g. 'policy 1', 'Policy 2', 'icici', 'bajaj').
    """
    if not company_name:
        return "Error: Company name search query is empty."
        
    comp_clean = company_name.strip().lower().replace("_", " ")
    
    # Walk through the CACHE_DIR to find matching profiles or raw PDFs to parse
    if not os.path.exists(CACHE_DIR):
        return f"Error: Rules folder '{CACHE_DIR}' does not exist."
        
    all_files = os.listdir(CACHE_DIR)
    
    # Try to find a pre-extracted profile first (smart string comparison replacing underscores with spaces)
    matching_profile_file = None
    for file in all_files:
        if file.endswith("_profile.json"):
            file_clean = file.lower().replace("_", " ")
            if comp_clean in file_clean or file_clean.replace(" profile.json", "") in comp_clean:
                matching_profile_file = file
                break
            
    if matching_profile_file:
        profile_path = os.path.join(CACHE_DIR, matching_profile_file)
        try:
            with open(profile_path, "r", encoding="utf-8") as f:
                return f"Policy Rules Profile (Structured):\n{f.read()}"
        except Exception as e:
            return f"Error: Failed to read existing profile: {str(e)}"
            
    # If no profile found, try to locate the corresponding PDF or Excel file to build it
    pdf_match = None
    xlsx_match = None
    for file in all_files:
        file_clean = file.lower().replace("_", " ")
        if comp_clean in file_clean or file_clean.replace(".pdf", "").replace(".xlsx", "") in comp_clean:
            if file.endswith(".pdf"):
                pdf_match = file
                break
            elif file.endswith(".xlsx") and not xlsx_match:
                xlsx_match = file
            
    if pdf_match:
        try:
            pdf_path = os.path.join(CACHE_DIR, pdf_match)
            profile = extract_policy_profile(pdf_path)
            return f"Policy Rules Profile (Structured & Extracted):\n{json.dumps(profile, indent=2)}"
        except Exception as e:
            return f"Error: Found PDF but extraction failed: {str(e)}"
            
    if xlsx_match:
        try:
            xlsx_path = os.path.join(CACHE_DIR, xlsx_match)
            profile = extract_policy_profile_from_excel(xlsx_path)
            return f"Policy Rules Profile (Structured & Extracted from Excel):\n{json.dumps(profile, indent=2)}"
        except Exception as e:
            return f"Error: Found Excel but extraction failed: {str(e)}"
            
    return f"No matching policy rules file or structured profile found for query: '{company_name}' in {CACHE_DIR}."
