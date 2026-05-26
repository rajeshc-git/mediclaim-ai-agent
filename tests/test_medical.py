import unittest
import os
import sys
import json
import shutil

# Ensure agent module is discoverable by test path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from agent.tools.medical import search_medical_codes, check_policy_limit, validate_claim_form, calculate_claim_probability
from agent.tools.file_ops import list_directory_recursive, read_pdf, read_excel

class TestMedicalTools(unittest.TestCase):
    """
    Unit test cases for the clinical coding index, policy rules calculator,
    claims form validator, and recursive directory tree listing.
    """
    def setUp(self):
        # Setup temporary directories for testing file operations
        self.test_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "test_workspace"))
        os.makedirs(self.test_dir, exist_ok=True)
        
        # Create nested folders
        self.sub_dir = os.path.join(self.test_dir, "patient_records", "john_doe")
        os.makedirs(self.sub_dir, exist_ok=True)
        
        # Create a text file
        self.txt_file = os.path.join(self.sub_dir, "clinical_notes.txt")
        with open(self.txt_file, "w", encoding="utf-8") as f:
            f.write("Patient is suffering from acute low back pain.")

    def tearDown(self):
        # Clean up temporary test files
        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)
            
        # Clean up claims_audit test logs
        claims_dir = os.path.normpath("claims_audit")
        test_claim_json = os.path.join(claims_dir, "claim_TEST999.json")
        test_claim_xlsx = os.path.join(claims_dir, "claim_TEST999.xlsx")
        if os.path.exists(test_claim_json):
            os.remove(test_claim_json)
        if os.path.exists(test_claim_xlsx):
            os.remove(test_claim_xlsx)

    def test_search_medical_codes(self):
        # Test exact code matching
        res_icd = search_medical_codes("M54.5")
        self.assertIn("Low back pain", res_icd)
        
        # Test keyword matching (case-insensitive)
        res_cpt = search_medical_codes("asthma")
        self.assertIn("J45.909", res_cpt)
        
        # Test non-existent code
        res_empty = search_medical_codes("nonexistentcode")
        self.assertIn("No medical codes", res_empty)

    def test_check_policy_limit(self):
        # Test Gold Tier outpatient
        res_gold = check_policy_limit("gold", "outpatient")
        self.assertIn("GOLD", res_gold)
        self.assertIn("90%", res_gold)
        self.assertIn("10%", res_gold)  # patient co-insurance
        
        # Test Silver Tier cardiac care
        res_silver = check_policy_limit("silver", "cardiac_care")
        self.assertIn("SILVER", res_silver)
        self.assertIn("80%", res_silver)
        
        # Test invalid inputs
        res_err_tier = check_policy_limit("platinum", "outpatient")
        self.assertTrue(res_err_tier.startswith("Error"))
        
        res_err_claim = check_policy_limit("gold", "surgery_unknown")
        self.assertTrue(res_err_claim.startswith("Error"))

    def test_validate_claim_form(self):
        # Test valid claim calculation
        res = validate_claim_form("TEST999", "John Doe", 1000.00, 800.00)
        self.assertIn("recorded successfully", res)
        self.assertIn("claims_audit/claim_TEST999.json", res)
        self.assertIn("claims_audit/claim_TEST999.xlsx", res)
        
        # Verify JSON file creation
        claims_dir = os.path.normpath("claims_audit")
        filepath = os.path.join(claims_dir, "claim_TEST999.json")
        self.assertTrue(os.path.exists(filepath))
        
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
            self.assertEqual(data["claim_id"], "TEST999")
            self.assertEqual(data["patient_copay_responsibility_inr"], 200.00)
            
        # Test validation error check
        res_err = validate_claim_form("TEST999", "John Doe", 500.00, 600.00)
        self.assertTrue(res_err.startswith("Error"))

    def test_validate_claim_form_multi_policy(self):
        multi_data = [
            {
                "company_name": "ICICI Lombard",
                "policy_tier": "Gold",
                "approved_coverage": 8500.0,
                "patient_copay": 1500.0,
                "probability": 80,
                "risk_category": "MODERATE RISK",
                "verdict": "APPROVED WITH ROOM CAPPING",
                "deductions": ["Room Rent Capping Violation (-20%)"],
                "justification": "Billed room rent exceeds the policy's 1% normal sum insured limit."
            },
            {
                "company_name": "Bajaj Allianz",
                "policy_tier": "Silver",
                "approved_coverage": 9000.0,
                "patient_copay": 1000.0,
                "probability": 90,
                "risk_category": "LOW RISK",
                "verdict": "RECOMMENDED FOR INSTANT CASHLESS APPROVAL",
                "deductions": [],
                "justification": "Compliant claim within Bajaj's policy limits."
            }
        ]
        
        multi_data_str = json.dumps(multi_data)
        
        res = validate_claim_form("TEST999", "John Doe", 10000.00, 9000.00, multi_policy_audit_data=multi_data_str)
        self.assertIn("recorded successfully", res)
        self.assertIn("Multi-Policy Comparison Summary", res)
        self.assertIn("ICICI Lombard", res)
        self.assertIn("Bajaj Allianz", res)
        
        # Verify JSON file creation
        claims_dir = os.path.normpath("claims_audit")
        filepath_json = os.path.join(claims_dir, "claim_TEST999.json")
        self.assertTrue(os.path.exists(filepath_json))
        
        with open(filepath_json, "r", encoding="utf-8") as f:
            data = json.load(f)
            self.assertEqual(data["claim_id"], "TEST999")
            self.assertIn("multi_policy_audits", data)
            self.assertEqual(len(data["multi_policy_audits"]), 2)
            self.assertEqual(data["multi_policy_audits"][0]["company_name"], "ICICI Lombard")
            
        # Verify Excel file creation and that it has multiple sheets
        filepath_xlsx = os.path.join(claims_dir, "claim_TEST999.xlsx")
        self.assertTrue(os.path.exists(filepath_xlsx))
        
        import openpyxl
        wb = openpyxl.load_workbook(filepath_xlsx)
        self.assertIn("Audit Dashboard", wb.sheetnames)
        self.assertIn("ICICI Lombard", wb.sheetnames)
        self.assertIn("Bajaj Allianz", wb.sheetnames)

    def test_list_directory_recursive(self):
        res = list_directory_recursive(self.test_dir)
        self.assertIn("john_doe", res)
        self.assertIn("clinical_notes.txt", res)

    def test_read_pdf_not_found(self):
        res = read_pdf("nonexistent.pdf")
        self.assertTrue(res.startswith("Error"))

    def test_read_excel(self):
        # Test nonexistent file
        res = read_excel("nonexistent.xlsx")
        self.assertTrue(res.startswith("Error"))
        
        # Test real file sheets listing
        real_file = r"insurance_rules/Aditya_Birla_Insurance_Master_6789_Data.xlsx"
        if os.path.exists(real_file):
            res_sheets = read_excel(real_file)
            self.assertIn("BED & ROOM CHARGES", res_sheets)
            self.assertIn("SURGICAL PROCEDURES", res_sheets)
            
            # Test smart resolution of bare filename
            res_smart = read_excel("Aditya_Birla_Insurance_Master_6789_Data.xlsx")
            self.assertIn("BED & ROOM CHARGES", res_smart)
            
            # Test reading specific sheet
            res_data = read_excel(real_file, "BED & ROOM CHARGES")
            self.assertIn("IPD BED CHARGES", res_data)
            self.assertIn("| ROOM CATEGORY | Rate (₹/day) |", res_data)

    def test_calculate_claim_probability(self):
        # Test perfect claim score
        res_perfect = calculate_claim_probability("Bajaj Allianz")
        self.assertIn("100%", res_perfect)
        self.assertIn("Bajaj Allianz Health Insurance", res_perfect)
        self.assertIn("LOW RISK", res_perfect)
        
        # Test claim with multiple violations
        res_violations = calculate_claim_probability(
            provider_name="ICICI Lombard",
            room_rent_exceeded=True,
            ped_detected=True,
            coding_mismatch=True
        )
        # 100 - 20 - 35 - 15 = 30%
        self.assertIn("30%", res_violations)
        self.assertIn("ICICI Lombard General Insurance", res_violations)
        self.assertIn("HIGH RISK", res_violations)

if __name__ == "__main__":
    unittest.main()
