import os
from agent.tools.base import tool

def resolve_smart_path(input_path: str) -> str:
    """
    Attempts to resolve the input_path. If the file exists directly, returns it normalized.
    Otherwise, if the file doesn't exist, extracts the filename (basename) and recursively
    searches the workspace for a matching file, skipping known system and temporary folders.
    """
    safe_path = os.path.normpath(input_path)
    if os.path.exists(safe_path):
        return safe_path
        
    filename = os.path.basename(input_path)
    if not filename:
        return safe_path
        
    # Search recursively in the current working directory
    workspace_root = os.getcwd()
    for root, dirs, files in os.walk(workspace_root):
        # Skip system, git, python env, and temporary testing directories to ensure safety and speed
        dirs[:] = [
            d for d in dirs
            if d not in (
                ".git", ".pytest_cache", ".vscode", "node_modules", 
                "__pycache__", "venv", ".venv", "test_workspace"
            )
        ]
        
        if filename in files:
            matched_path = os.path.join(root, filename)
            return os.path.normpath(matched_path)
            
    return safe_path

@tool
def read_file(filename: str) -> str:
    """
    Reads the content of a file from the local workspace.
    
    Args:
        filename: The absolute or relative path to the file to read.
    """
    # Safe path resolution - keep operations within workspace
    safe_path = resolve_smart_path(filename)
    if not os.path.exists(safe_path):
        return f"Error: File '{filename}' does not exist."
    
    if os.path.isdir(safe_path):
        return f"Error: '{filename}' is a directory, not a file. Use list_directory instead."
        
    try:
        with open(safe_path, "r", encoding="utf-8") as f:
            content = f.read()
        return content
    except Exception as e:
        return f"Error: Failed to read file: {str(e)}"

@tool
def write_file(filename: str, content: str) -> str:
    """
    Creates or overwrites a file with the specified content.
    
    Args:
        filename: The path where the file will be saved.
        content: The text content to write into the file.
    """
    safe_path = os.path.normpath(filename)
    # Ensure any parent directories exist
    parent_dir = os.path.dirname(safe_path)
    if parent_dir and not os.path.exists(parent_dir):
        try:
            os.makedirs(parent_dir, exist_ok=True)
        except Exception as e:
            return f"Error: Failed to create directories for path: {str(e)}"

    try:
        with open(safe_path, "w", encoding="utf-8") as f:
            f.write(content)
        size = os.path.getsize(safe_path)
        return f"Success: Wrote file successfully to '{filename}' ({size} bytes)."
    except Exception as e:
        return f"Error: Failed to write file: {str(e)}"

@tool
def list_directory(directory_path: str = ".") -> str:
    """
    Lists the files and folders inside the specified directory path.
    Files are returned sorted by their last modified/uploaded timestamp (newest first).
    
    Args:
        directory_path: The directory path to list. Defaults to the current directory.
    """
    import datetime
    safe_path = os.path.normpath(directory_path)
    if not os.path.exists(safe_path):
        return f"Error: Directory '{directory_path}' does not exist."
        
    if not os.path.isdir(safe_path):
        return f"Error: '{directory_path}' is a file, not a directory."
        
    try:
        items = os.listdir(safe_path)
        if not items:
            return f"The directory '{directory_path}' is empty."
            
        # Separate files and folders to sort files by mtime
        folders_list = []
        files_list = []
        
        for item in items:
            item_path = os.path.join(safe_path, item)
            mtime = os.path.getmtime(item_path)
            mtime_str = datetime.datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
            
            if os.path.isdir(item_path):
                folders_list.append((item, mtime_str))
            else:
                size = os.path.getsize(item_path)
                files_list.append((item, size, mtime, mtime_str))
                
        # Sort files by modification time (newest first)
        files_list.sort(key=lambda x: x[2], reverse=True)
        
        formatted_list = []
        for folder, mtime_str in folders_list:
            formatted_list.append(f"- 📁 {folder}/ (Modified: {mtime_str})")
            
        for file, size, _, mtime_str in files_list:
            formatted_list.append(f"- 📄 {file} ({size} bytes) (Uploaded/Modified: {mtime_str})")
            
        return "\n".join(formatted_list)
    except Exception as e:
        return f"Error: Failed to list directory: {str(e)}"

@tool
def read_pdf(filepath: str) -> str:
    """
    Extracts text content from a PDF file located at the specified workspace path.
    
    Args:
        filepath: The absolute or relative path to the PDF file to read.
    """
    safe_path = resolve_smart_path(filepath)
    if not os.path.exists(safe_path):
        return f"Error: PDF file '{filepath}' does not exist."
        
    if os.path.isdir(safe_path):
        return f"Error: '{filepath}' is a directory, not a file."
        
    try:
        # pyrefly: ignore [missing-import]
        from pypdf import PdfReader
        reader = PdfReader(safe_path)
        
        extracted_text = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                extracted_text.append(f"--- PAGE {i+1} ---\n{text.strip()}")
                
        if not extracted_text:
            return f"Warning: PDF file '{filepath}' was successfully read but contained no extractable text."
            
        return "\n\n".join(extracted_text)
    except Exception as e:
        return f"Error: Failed to parse PDF: {str(e)}"

@tool
def list_directory_recursive(directory_path: str = ".") -> str:
    """
    Recursively lists all files and folders inside the specified directory path.
    All files across folders are returned with sizes and last modified timestamps.
    
    Args:
        directory_path: The root directory path to scan. Defaults to the current directory.
    """
    import datetime
    safe_path = os.path.normpath(directory_path)
    if not os.path.exists(safe_path):
        return f"Error: Target path '{directory_path}' does not exist."
        
    if not os.path.isdir(safe_path):
        return f"Error: '{directory_path}' is a file, not a directory."
        
    try:
        all_items = []
        for root, dirs, files in os.walk(safe_path):
            for file in files:
                file_path = os.path.join(root, file)
                size = os.path.getsize(file_path)
                mtime = os.path.getmtime(file_path)
                mtime_str = datetime.datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
                rel_dir = os.path.relpath(root, safe_path)
                all_items.append({
                    "name": file,
                    "rel_dir": rel_dir,
                    "size": size,
                    "mtime": mtime,
                    "mtime_str": mtime_str,
                    "full_path": file_path
                })
                
        if not all_items:
            return f"The directory '{directory_path}' contains no files."
            
        # Sort items globally by modification time (newest first) so recent uploads stand out
        all_items.sort(key=lambda x: x["mtime"], reverse=True)
        
        formatted_tree = [f"📁 {os.path.basename(os.path.abspath(safe_path)) or directory_path}/ (Files ordered by newest upload/modified time):"]
        for item in all_items:
            dir_prefix = "" if item["rel_dir"] == "." else f"{item['rel_dir']}/"
            formatted_tree.append(f"  📄 {dir_prefix}{item['name']} ({item['size']} bytes) [Uploaded: {item['mtime_str']}]")
            
        return "\n".join(formatted_tree)
    except Exception as e:
        return f"Error: Failed to recursively walk directory: {str(e)}"

@tool
def read_excel(filepath: str, sheet_name: str = "") -> str:
    """
    Reads the content of an Excel spreadsheet (.xlsx) and presents it as a beautifully formatted Markdown table.
    Use this to inspect exact numerical tables, charges, surgical package rates, and diagnostic fees.
    If sheet_name is not provided, it lists all available sheet names in the workbook.
    
    Args:
        filepath: The absolute or relative path to the Excel file (.xlsx) to read.
        sheet_name: Optional name of the specific worksheet to load (case-sensitive). If empty, returns available sheets.
    """
    import openpyxl
    
    safe_path = resolve_smart_path(filepath)
    if not os.path.exists(safe_path):
        return f"Error: Excel file '{filepath}' does not exist."
        
    if os.path.isdir(safe_path):
        return f"Error: '{filepath}' is a directory, not a file."
        
    try:
        wb = openpyxl.load_workbook(safe_path, data_only=True)
        if not sheet_name:
            sheets = wb.sheetnames
            return f"Workbook sheets in '{filepath}':\n" + "\n".join(f"- {s}" for s in sheets) + "\n\nSpecify one of these sheet names to read the data."
            
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            
        ws = wb[sheet_name]
        markdown_lines = []
        
        # Read all rows
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return f"The sheet '{sheet_name}' is empty."
            
        # Clean rows (remove trailing None-only rows/columns to make output clean)
        cleaned_rows = []
        for row in rows:
            row_list = list(row)
            while row_list and row_list[-1] is None:
                row_list.pop()
            if any(x is not None for x in row_list):
                cleaned_rows.append(row_list)
                
        if not cleaned_rows:
            return f"The sheet '{sheet_name}' contains no readable data."
            
        # Find maximum columns
        max_cols = max(len(r) for r in cleaned_rows)
        
        # Format rows as Markdown table
        for idx, row in enumerate(cleaned_rows):
            # Pad row if shorter than max_cols
            padded_row = [str(x) if x is not None else "" for x in row] + [""] * (max_cols - len(row))
            # Format cell values to keep visual alignment clean and prevent multi-line breaks
            padded_row = [x.replace("\n", " ").replace("|", "\\|") for x in padded_row]
            
            row_str = "| " + " | ".join(padded_row) + " |"
            markdown_lines.append(row_str)
            
            # Add table separator below header/first row
            if idx == 0:
                separator = "| " + " | ".join(["---"] * max_cols) + " |"
                markdown_lines.append(separator)
                
        return f"### Excel Data Sheet: {sheet_name} ({len(cleaned_rows)} rows)\n\n" + "\n".join(markdown_lines)
    except Exception as e:
        return f"Error: Failed to read Excel workbook: {str(e)}"

